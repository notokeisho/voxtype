"""Admin global dictionary management API endpoints."""

from datetime import datetime
import io

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, status
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict
from sqlalchemy import delete, select

from app.auth.dependencies import get_current_admin_user
from app.database import async_session_factory
from app.models.global_dictionary import GlobalDictionary
from app.models.user import User

router = APIRouter(prefix="/admin/api", tags=["admin"])


class DictionaryEntryResponse(BaseModel):
    """Dictionary entry response model."""

    model_config = ConfigDict(from_attributes=True)

    id: int
    pattern: str
    replacement: str
    created_at: datetime
    created_by: int | None


class AddDictionaryEntryRequest(BaseModel):
    """Request model for adding dictionary entry."""

    pattern: str
    replacement: str


class ImportDictionaryResponse(BaseModel):
    """Response for importing dictionary entries."""

    added: int
    skipped: int
    failed: int


@router.get("/dictionary", response_model=list[DictionaryEntryResponse])
async def list_global_dictionary(
    _admin: User = Depends(get_current_admin_user),
) -> list[DictionaryEntryResponse]:
    """List all global dictionary entries.

    Admin only endpoint.
    """
    async with async_session_factory() as session:
        result = await session.execute(
            select(GlobalDictionary).order_by(GlobalDictionary.created_at.desc())
        )
        entries = result.scalars().all()
        return [DictionaryEntryResponse.model_validate(e) for e in entries]


@router.post(
    "/dictionary",
    response_model=DictionaryEntryResponse,
    status_code=status.HTTP_201_CREATED,
)
async def add_global_dictionary_entry(
    request: AddDictionaryEntryRequest,
    admin: User = Depends(get_current_admin_user),
) -> DictionaryEntryResponse:
    """Add a global dictionary entry.

    Admin only endpoint.
    """
    async with async_session_factory() as session:
        # Check if pattern already exists
        result = await session.execute(
            select(GlobalDictionary).where(
                GlobalDictionary.pattern == request.pattern
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Pattern already exists in global dictionary",
            )

        entry = GlobalDictionary(
            pattern=request.pattern,
            replacement=request.replacement,
            created_by=admin.id,
        )
        session.add(entry)
        await session.commit()
        await session.refresh(entry)

        return DictionaryEntryResponse.model_validate(entry)


@router.delete("/dictionary/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_global_dictionary_entry(
    entry_id: int,
    _admin: User = Depends(get_current_admin_user),
) -> None:
    """Delete a global dictionary entry.

    Admin only endpoint.
    """
    async with async_session_factory() as session:
        result = await session.execute(
            select(GlobalDictionary).where(GlobalDictionary.id == entry_id)
        )
        entry = result.scalar_one_or_none()

        if entry is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Dictionary entry not found",
            )

        await session.execute(
            delete(GlobalDictionary).where(GlobalDictionary.id == entry_id)
        )
        await session.commit()


@router.get("/dictionary/export")
async def export_global_dictionary_xlsx(
    _admin: User = Depends(get_current_admin_user),
) -> Response:
    """Export global dictionary as XLSX."""
    async with async_session_factory() as session:
        result = await session.execute(
            select(GlobalDictionary).order_by(GlobalDictionary.created_at.desc())
        )
        entries = result.scalars().all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "global_dictionary"
    sheet.append(["pattern", "replacement", "created_at", "created_by"])
    for entry in entries:
        created_at = entry.created_at.isoformat() if entry.created_at else ""
        sheet.append([entry.pattern, entry.replacement, created_at, entry.created_by])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "global_dictionary.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/dictionary/import", response_model=ImportDictionaryResponse)
async def import_global_dictionary_xlsx(
    file: UploadFile,
    _admin: User = Depends(get_current_admin_user),
) -> ImportDictionaryResponse:
    """Import global dictionary entries from XLSX."""
    if file.filename is None or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format",
        )

    try:
        content = await file.read()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to read XLSX file",
        ) from exc

    # Load workbook from bytes
    try:
        workbook = load_workbook(io.BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid XLSX file",
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="XLSX file is empty",
        )

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    expected_header = ["pattern", "replacement", "created_at", "created_by"]
    if header != expected_header:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid XLSX header",
        )

    added = 0
    skipped = 0
    failed = 0
    processed = 0

    async with async_session_factory() as session:
        result = await session.execute(select(GlobalDictionary.pattern))
        existing_patterns = {row[0] for row in result.all()}

        for row in rows[1:]:
            if row is None:
                continue
            pattern = str(row[0]).strip() if row[0] is not None else ""
            replacement = str(row[1]).strip() if row[1] is not None else ""

            if not pattern and not replacement:
                continue

            if not pattern or not replacement:
                failed += 1
                break

            processed += 1
            if processed > 10000:
                failed += 1
                break

            if pattern in existing_patterns:
                skipped += 1
                continue

            entry = GlobalDictionary(pattern=pattern, replacement=replacement, created_by=None)
            session.add(entry)
            existing_patterns.add(pattern)
            added += 1

        if failed > 0:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLSX rows",
            )

        await session.commit()

    return ImportDictionaryResponse(added=added, skipped=skipped, failed=failed)
