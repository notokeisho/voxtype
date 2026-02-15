"""Backup restore service utilities."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.models.global_dictionary import GlobalDictionary


@dataclass(slots=True)
class BackupRestoreResult:
    """Result of restoring dictionary from backup."""

    restored_file: str
    mode: str
    total: int
    added: int
    skipped: int
    failed: int
    restored_at: datetime


def _validate_header(header_row: tuple) -> bool:
    header = [str(cell).strip() if cell is not None else "" for cell in header_row]
    return header == ["pattern", "replacement", "created_at", "created_by"]


async def restore_global_dictionary_from_backup(
    session,
    filename: str,
    mode: str,
    base_dir: Path | None = None,
) -> BackupRestoreResult:
    """Restore global dictionary from backup file."""
    backup_dir = base_dir or Path("./data/backups")
    backup_file = backup_dir / filename
    if not backup_file.exists():
        raise FileNotFoundError(filename)

    workbook = load_workbook(backup_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Backup file is empty")
    if not _validate_header(rows[0]):
        raise ValueError("Invalid backup header")

    if mode == "replace":
        await session.execute(delete(GlobalDictionary))
        existing_patterns = set()
    else:
        result = await session.execute(select(GlobalDictionary.pattern))
        existing_patterns = {row[0] for row in result.all()}

    total = 0
    added = 0
    skipped = 0
    failed = 0

    for row in rows[1:]:
        if row is None:
            continue
        pattern = str(row[0]).strip() if row[0] is not None else ""
        replacement = str(row[1]).strip() if row[1] is not None else ""
        if not pattern and not replacement:
            continue

        total += 1
        if not pattern or not replacement:
            failed += 1
            continue

        if pattern in existing_patterns:
            skipped += 1
            continue

        session.add(
            GlobalDictionary(
                pattern=pattern,
                replacement=replacement,
                created_by=None,
            )
        )
        existing_patterns.add(pattern)
        added += 1

    await session.commit()

    return BackupRestoreResult(
        restored_file=backup_file.name,
        mode=mode,
        total=total,
        added=added,
        skipped=skipped,
        failed=failed,
        restored_at=datetime.now(),
    )
