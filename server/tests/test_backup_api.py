"""Tests for backup admin API endpoints."""

from pathlib import Path
import asyncio

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook
from sqlalchemy import delete, select, text

from app.auth.jwt import create_jwt_token
from app.database import async_session_factory
from app.main import app
from app.models.global_dictionary import GlobalDictionary
from app.models.user import User
from app.models.whitelist import Whitelist


@pytest.fixture
async def admin_user():
    """Create an admin user."""
    async with async_session_factory() as session:
        await session.execute(text("DELETE FROM backup_settings"))
        await session.execute(delete(Whitelist).where(Whitelist.github_id == "backupadmin"))
        await session.execute(delete(User).where(User.github_id == "backupadmin"))
        await session.commit()

        user = User(
            github_id="backupadmin",
            github_avatar="https://example.com/admin.png",
            is_admin=True,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)

        whitelist_entry = Whitelist(github_id="backupadmin")
        session.add(whitelist_entry)
        await session.commit()

        yield user

        await session.execute(delete(Whitelist).where(Whitelist.github_id == "backupadmin"))
        await session.execute(delete(User).where(User.github_id == "backupadmin"))
        await session.execute(text("DELETE FROM backup_settings"))
        await session.commit()


@pytest.fixture
def admin_token(admin_user):
    """Create JWT token for admin user."""
    return create_jwt_token(user_id=admin_user.id, github_id=admin_user.github_id)


class TestBackupSettingsApi:
    """Tests for backup settings endpoints."""

    @pytest.mark.asyncio
    async def test_get_backup_settings_default(self, admin_token):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.get(
                "/admin/api/dictionary/backup",
                headers={"Authorization": f"Bearer {admin_token}"},
            )

        assert response.status_code == 200
        data = response.json()
        assert data["enabled"] is False
        assert data["last_run_at"] is None

    @pytest.mark.asyncio
    async def test_update_backup_settings(self, admin_token):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.patch(
                "/admin/api/dictionary/backup",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={"enabled": True},
            )

        assert response.status_code == 200
        data = response.json()
        assert data["enabled"] is True


class TestBackupRunApi:
    """Tests for backup run endpoint."""

    @pytest.mark.asyncio
    async def test_run_backup(self, admin_token):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/admin/api/dictionary/backup/run",
                headers={"Authorization": f"Bearer {admin_token}"},
            )

        assert response.status_code == 200
        data = response.json()
        assert data["created_file"].endswith(".xlsx")
        assert data["created_at"] is not None
        assert isinstance(data["kept"], int)
        assert isinstance(data["deleted"], int)


class TestBackupFilesApi:
    """Tests for backup files listing endpoint."""

    @pytest.mark.asyncio
    async def test_list_backup_files_returns_sorted_files(self, admin_token):
        backup_dir = Path("./data/backups")
        backup_dir.mkdir(parents=True, exist_ok=True)
        target_files = [
            backup_dir / "global_dictionary_2026-02-15_12-00-00.xlsx",
            backup_dir / "global_dictionary_2026-02-14_12-00-00.xlsx",
        ]
        for file_path in target_files:
            file_path.write_bytes(b"test")

        try:
            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.get(
                    "/admin/api/dictionary/backup/files",
                    headers={"Authorization": f"Bearer {admin_token}"},
                )
        finally:
            for file_path in target_files:
                file_path.unlink(missing_ok=True)

        assert response.status_code == 200
        data = response.json()
        assert "files" in data

        listed_names = [item["filename"] for item in data["files"]]
        assert target_files[0].name in listed_names
        assert target_files[1].name in listed_names
        assert listed_names.index(target_files[0].name) < listed_names.index(target_files[1].name)

        first_item = data["files"][0]
        assert "filename" in first_item
        assert "created_at" in first_item
        assert "size_bytes" in first_item


class TestBackupRestoreApi:
    """Tests for backup restore endpoint."""

    @pytest.mark.asyncio
    async def test_restore_backup_merge_mode(self, admin_token):
        backup_dir = Path("./data/backups")
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_file = backup_dir / "global_dictionary_2026-02-16_03-00-00.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "global_dictionary"
        sheet.append(["pattern", "replacement", "created_at", "created_by"])
        sheet.append(["restore_task2_existing", "new_value", "", ""])
        sheet.append(["restore_task2_added", "added_value", "", ""])
        workbook.save(backup_file)

        async with async_session_factory() as session:
            session.add(
                GlobalDictionary(
                    pattern="restore_task2_existing",
                    replacement="old_value",
                    created_by=None,
                )
            )
            await session.commit()

        try:
            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.post(
                    "/admin/api/dictionary/backup/restore",
                    headers={"Authorization": f"Bearer {admin_token}"},
                    json={"filename": backup_file.name, "mode": "merge"},
                )
        finally:
            backup_file.unlink(missing_ok=True)
            async with async_session_factory() as session:
                await session.execute(
                    delete(GlobalDictionary).where(
                        GlobalDictionary.pattern.in_(
                            ["restore_task2_existing", "restore_task2_added"]
                        )
                    )
                )
                await session.commit()

        assert response.status_code == 200
        data = response.json()
        assert data["mode"] == "merge"
        assert data["restored_file"] == backup_file.name
        assert data["added"] == 1
        assert data["skipped"] == 1
        assert data["failed"] == 0
        assert data["total"] == 2

    @pytest.mark.asyncio
    async def test_restore_backup_replace_mode(self, admin_token):
        backup_dir = Path("./data/backups")
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_file = backup_dir / "global_dictionary_2026-02-16_03-10-00.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "global_dictionary"
        sheet.append(["pattern", "replacement", "created_at", "created_by"])
        sheet.append(["restore_task3_replace_only", "replace_value", "", ""])
        workbook.save(backup_file)

        async with async_session_factory() as session:
            session.add(
                GlobalDictionary(
                    pattern="restore_task3_should_be_removed",
                    replacement="old_value",
                    created_by=None,
                )
            )
            await session.commit()

        try:
            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.post(
                    "/admin/api/dictionary/backup/restore",
                    headers={"Authorization": f"Bearer {admin_token}"},
                    json={"filename": backup_file.name, "mode": "replace"},
                )

            async with async_session_factory() as session:
                result = await session.execute(select(GlobalDictionary.pattern))
                patterns = {row[0] for row in result.all()}
        finally:
            backup_file.unlink(missing_ok=True)
            async with async_session_factory() as session:
                await session.execute(
                    delete(GlobalDictionary).where(
                        GlobalDictionary.pattern.in_(
                            [
                                "restore_task3_replace_only",
                                "restore_task3_should_be_removed",
                            ]
                        )
                    )
                )
                await session.commit()

        assert response.status_code == 200
        data = response.json()
        assert data["mode"] == "replace"
        assert data["restored_file"] == backup_file.name
        assert data["added"] == 1
        assert data["skipped"] == 0
        assert data["failed"] == 0
        assert data["total"] == 1
        assert "restore_task3_replace_only" in patterns
        assert "restore_task3_should_be_removed" not in patterns

    @pytest.mark.asyncio
    async def test_restore_backup_rejects_invalid_filename(self, admin_token):
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/admin/api/dictionary/backup/restore",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={"filename": "../secrets.txt", "mode": "merge"},
            )

        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_restore_backup_returns_conflict_when_locked(self, admin_token):
        backup_dir = Path("./data/backups")
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_file = backup_dir / "global_dictionary_2026-02-16_03-20-00.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "global_dictionary"
        sheet.append(["pattern", "replacement", "created_at", "created_by"])
        sheet.append(["restore_task4_conflict", "value", "", ""])
        workbook.save(backup_file)

        async def call_restore(client: AsyncClient):
            return await client.post(
                "/admin/api/dictionary/backup/restore",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={"filename": backup_file.name, "mode": "merge"},
            )

        try:
            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                first, second = await asyncio.gather(call_restore(client), call_restore(client))
        finally:
            backup_file.unlink(missing_ok=True)
            async with async_session_factory() as session:
                await session.execute(
                    delete(GlobalDictionary).where(GlobalDictionary.pattern == "restore_task4_conflict")
                )
                await session.commit()

        statuses = {first.status_code, second.status_code}
        assert statuses == {200, 409}
