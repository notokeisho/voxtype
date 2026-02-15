"""Tests for backup generation."""

from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import settings
from app.models.global_dictionary import GlobalDictionary
from app.services import backup as backup_service
from app.services.backup import create_global_dictionary_backup


@dataclass
class BackupTestEntry:
    pattern: str
    replacement: str


def run_async(coro):
    """Run async coroutine in a new event loop."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


async def _with_session(callback):
    engine = create_async_engine(settings.database_url)
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        async with async_session() as session:
            return await callback(session)
    finally:
        await engine.dispose()


async def _cleanup_entries(prefix: str) -> None:
    async def _cleanup(session: AsyncSession) -> None:
        await session.execute(
            text("DELETE FROM global_dictionary WHERE pattern LIKE :prefix"),
            {"prefix": f"{prefix}%"},
        )
        await session.commit()

    await _with_session(_cleanup)


async def _insert_entries(entries: list[BackupTestEntry]) -> None:
    async def _insert(session: AsyncSession) -> None:
        for entry in entries:
            session.add(
                GlobalDictionary(
                    pattern=entry.pattern,
                    replacement=entry.replacement,
                    created_by=None,
                )
            )
        await session.commit()

    await _with_session(_insert)


async def _create_backup(
    backup_dir: Path,
    target_date: date,
):
    async def _create(session: AsyncSession):
        return await create_global_dictionary_backup(
            session,
            base_dir=backup_dir,
            current_date=target_date,
            now_provider=lambda: datetime(2026, 2, 15, 0, 0, 0),
        )

    return await _with_session(_create)


def test_backup_creates_xlsx_with_entries(tmp_path: Path):
    prefix = "backup_test_entries_"
    entries = [
        BackupTestEntry(pattern=f"{prefix}one", replacement="ONE"),
        BackupTestEntry(pattern=f"{prefix}two", replacement="TWO"),
    ]

    try:
        run_async(_cleanup_entries(prefix))
        run_async(_insert_entries(entries))
        result = run_async(_create_backup(tmp_path, date(2026, 2, 15)))

        assert result.created_file.name == "global_dictionary_2026-02-15_00-00-00.xlsx"
        assert result.created_file.exists()

        workbook = load_workbook(result.created_file)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0] == ("pattern", "replacement", "created_at", "created_by")
        patterns = {row[0] for row in rows[1:]}
        replacements = {row[1] for row in rows[1:]}
        assert f"{prefix}one" in patterns
        assert f"{prefix}two" in patterns
        assert "ONE" in replacements
        assert "TWO" in replacements
    finally:
        run_async(_cleanup_entries(prefix))


def test_backup_keeps_latest_three(tmp_path: Path):
    prefix = "backup_test_retention_"
    try:
        run_async(_cleanup_entries(prefix))
        run_async(_insert_entries([BackupTestEntry(pattern=f"{prefix}one", replacement="ONE")]))

        existing_dates = [
            date(2026, 2, 10),
            date(2026, 2, 11),
            date(2026, 2, 12),
            date(2026, 2, 13),
        ]
        for existing_date in existing_dates:
            filename = f"global_dictionary_{existing_date.isoformat()}_00-00-00.xlsx"
            (tmp_path / filename).write_text("dummy", encoding="utf-8")

        result = run_async(_create_backup(tmp_path, date(2026, 2, 14)))

        remaining = sorted(path.name for path in tmp_path.glob("global_dictionary_*.xlsx"))
        assert remaining == [
            "global_dictionary_2026-02-12_00-00-00.xlsx",
            "global_dictionary_2026-02-13_00-00-00.xlsx",
            "global_dictionary_2026-02-14_00-00-00.xlsx",
        ]
        assert result.kept == 3
        assert result.deleted == 2
    finally:
        run_async(_cleanup_entries(prefix))


def test_backup_handles_empty_dictionary(tmp_path: Path, monkeypatch):
    class FakeSession:
        async def commit(self):
            return None

    fake_settings = SimpleNamespace(last_run_at=None)

    async def fake_get_entries(_session):
        return []

    async def fake_get_settings(_session):
        return fake_settings

    monkeypatch.setattr(backup_service, "get_global_entries", fake_get_entries)
    monkeypatch.setattr(backup_service, "get_backup_settings", fake_get_settings)

    result = run_async(
        create_global_dictionary_backup(
            FakeSession(),
            base_dir=tmp_path,
            current_date=date(2026, 2, 15),
            now_provider=lambda: datetime(2026, 2, 15, 0, 0, 0),
        )
    )

    workbook = load_workbook(result.created_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == [("pattern", "replacement", "created_at", "created_by")]
